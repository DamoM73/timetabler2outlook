from openpyxl import load_workbook
from outlook_utils import Lesson

class Datastore:
    
    def __init__(self):
        self.file = "Sheet1.xlsx"
        self.workbook = load_workbook(filename = self.file)
        self.sheet = self.workbook.active
        self.periods = self.get_period_names()
        self.days = self.get_time_table()
                
                
    def get_value(self,cell):
        return self.sheet[cell].value
    
    
    def get_period_names(self):
        for row in self.sheet.iter_rows(min_row = 3,
                                       max_row = 3, 
                                       min_col = 2,
                                       max_col = 15,
                                       values_only = True):
            return(row)
    
    
    def get_time_table(self):
        days = []
        for row in self.sheet.iter_rows(min_row = 4,
                                       max_row = 8, 
                                       min_col = 2,
                                       max_col = 15,
                                       values_only = True):
            days.append(row)
            
        return days
    
    
    def get_lesson_and_room(self, raw):
        """
        extracts the lesson name and the room name from the value sent
        """
        if raw.find("\n") == -1:
            return(raw,None)
        else:
            name = raw[:raw.find("\n")]
            less_lesson = raw[raw.find("\n") + 1:]
            room = less_lesson[:less_lesson.find("\n")]
            return(name,room)
    
    
    def create_lessons(self):
        lessons = []
        day_names = ["Mon", "Tues", "Wed", "Thurs", "Fri"]
        lesson_times = [("",0),("08:00",20),("08:20",10,),("08:30",55),("09:25",55),("10:20",15),("10:35",15),
                        ("10:50",55),("11:45",55),("12:40",20),("13:00",20),("13:20",55),("14:15",55),("3:10",20)]

        for day_num, day in enumerate(self.days):
            for period_num, lesson in enumerate(day):                
                lesson_day = day_names[day_num]
                lesson_period = self.periods[period_num]
                lesson_time = lesson_times[period_num][0]
                lesson_duration = lesson_times[period_num][1]
                if lesson != "":
                    lesson_name, lesson_room = self.get_lesson_and_room(lesson)
                    if lesson_room == None:
                        lesson_room = lesson_name
                        lesson_name = "Duty"
                else:
                    lesson_name = ""
                    lesson_room = ""
                        
                lessons.append(Lesson(lesson_day, lesson_period, lesson_name, lesson_time, lesson_duration, lesson_room))
               
        return lessons